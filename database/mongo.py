import site
import pymongo
from pymongo import MongoClient
from openpyxl import load_workbook
import certifi


ca = certifi.where()
# from calculs.sponsor import sponsor

#ajoute une valeud dans la collection clients qui se trouve dans la base de donnees templates
def add_value_data_base(ticker, equity, inconvénient, dividende="", siteweb="", sponsor="", yahoo=""): 
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority", tlsCAFile=ca)
    
    db = cluster["templates"]
    collection = db["clients"]

    post1 = {"Ticker": ticker, "Equity" : equity, "Inconvénient": inconvénient, "Dividende": dividende, "SiteWeb": siteweb, "Sponsor": sponsor, "Yahoo": yahoo}

    collection.insert_one(post1)

# add_value_data_base("Cac50", "ca")


#montre les valeurs de la base de la collections clients de la base de données templates
def show_database():
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority", tlsCAFile=ca)

    db = cluster["templates"]
    collection = db["clients"]
    list_elements_database = []
    for x in collection.find({}, {"_id":0, "Libelle": 1, "Ticker": 1 }): 
        list_elements_database.append(x)

    return(list_elements_database)


# show_database()

def add_value():
    file_path = 'BDD_MongoDB_Projet_Brochure.xlsm'
    wb = load_workbook(file_path)
    ws = wb['Feuil1']  # or wb.active
    
    i = 2
    while True:
        ticker = "A" + str(i)
        Equity = "B" + str(i)
        inconvénient = "C" + str(i)
        dividende = "D" + str(i)
        siteweb = "E" + str(i)
        sponsor = "F" + str(i)
        yahoo = "G" + str(i)

        if (ws[ticker].value is None):
            exit (2)

        if (ws[ticker].value is None):   
            ws[ticker].value = " "
        
        add_value_data_base(ws[ticker].value, ws[Equity].value, ws[inconvénient].value, ws[dividende].value, ws[siteweb].value, ws[sponsor].value, ws[yahoo].value )
        i+=1
        
    wb.save(file_path)
# add_value()

def takeinformations(Class):
    #Recupere le sous jacent
    cluster = MongoClient("mongodb+srv://guillaume:guigui@cluster0.eczef.mongodb.net/myFirstDatabase?retryWrites=true&w=majority", tlsCAFile=ca)

    db = cluster["templates"]
    collection = db["clients"]

    #remplacer par le bon element(ici ticker)

    myresults = []
    #j ajoute les tickers
    for i in Class.TSJ:
        myresults.append(collection.find({"Ticker":i}))

    compteur = 0
    for result in myresults:
            if compteur == 0:
                mot = " "
                perf = " "
                
            if compteur == 1:
                mot = " et "
                perf = ", la performance positive ou négative de ce placement dépendant de l'évolution "
            
            if compteur > 2:
                mot = " et "
                perf = ""
            try:
                test = result[0]
            
                Class.NOMSOUSJACENT = Class.NOMSOUSJACENT + mot + (test["Equity"]) #on ajoute le sous jacent + ce qu on avait avant
                Class.DIVIDENDE = Class.DIVIDENDE + mot + test["Dividende"]
                Class.SPONSOR = Class.SPONSOR + mot + test["Sponsor"]
                Class.Site = Class.Site + mot + test["SiteWeb"]
                Class.TICKER = Class.TICKER + mot + test["Ticker"]
                
                try:
                    Class.inconvenient = str(Class.inconvenient) + mot + test["Inconvénient"]
                except Exception:
                    Class.inconvenient = str(Class.inconvenient) + mot + "NULL"

                Class.Yahoo.append(test["Yahoo"])
                Class.Yahoo_value_name.append(test["Equity"])
                Class.Yahoo_value_dividende.append(test["Dividende"])

            except Exception: #au cas ou si ca marche pas, pour éviter que ca crash
                    Class.NOMSOUSJACENT + mot + ("ERREUR LES POTES")
                    Class.DIVIDENDE = Class.DIVIDENDE + mot + "ERREUR"
                    Class.SPONSOR = Class.SPONSOR + mot + "ERREUR"
                    Class.Site = Class.Site + mot + "ERREUR"
                    Class.TICKER = Class.TICKER + mot + "ERREUR"
                    Class.Yahoo.append(test["Yahoo"])
                    
                    Class.Yahoo_value_name.append(test["Equity"])
                    Class.Yahoo_value_dividende.append("dividende inconnu")

                    Class.BLOCDIVIDENDE = "Le produit n'est pas identifié dans la Database"


            compteur+=1
        
    ticker = Class.Yahoo
    try:
        if (len(ticker == 1)):
            Class.BLOCDIVIDENDE = (f"{myresults[0][0]['Equity']} et {myresults[1][0]['Equity']}, la performance positive ou négative de ce placement dépendant de l'évolution de {Class.SJR7} {myresults[0][0]['Equity']} ({myresults[0][0]['Dividende']}; code Bloomberg : {myresults[0][0]['Ticker']}; place de cotation : {myresults[0][0]['Sponsor']}; {myresults[0][0]['SiteWeb']})")

        elif (len(ticker) == 2):
            Class.BLOCDIVIDENDE = (f"{myresults[0][0]['Equity']} et {myresults[1][0]['Equity']}, la performance positive ou négative de ce placement dépendant de l'évolution de {Class.SJR7} entre {myresults[0][0]['Equity']} ({myresults[0][0]['Dividende']}; code Bloomberg : {myresults[0][0]['Ticker']}; place de cotation : {myresults[0][0]['Sponsor']}; {myresults[0][0]['SiteWeb']}) et {myresults[1][0]['Equity']} ({myresults[1][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[1][0]['Sponsor']}; {myresults[1][0]['SiteWeb']})")

        elif (len(ticker) == 3):
            Class.BLOCDIVIDENDE = (f"{myresults[0][0]['Equity']} et {myresults[1][0]['Equity']}, la performance positive ou négative de ce placement dépendant de l'évolution de {Class.SJR7} entre {myresults[0][0]['Equity']} ({myresults[0][0]['Dividende']}; code Bloomberg : {myresults[0][0]['Ticker']}; place de cotation : {myresults[0][0]['Sponsor']}; {myresults[0][0]['SiteWeb']}) et {myresults[1][0]['Equity']} ({myresults[1][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[1][0]['Sponsor']}; {myresults[1][0]['SiteWeb']}) et {myresults[2][0]['Equity']} ({myresults[2][0]['Dividende']}; code Bloomberg : {myresults[2][0]['Ticker']}; place de cotation : {myresults[2][0]['Sponsor']}; {myresults[2][0]['SiteWeb']})")
        
        elif (len(ticker) == 4):
            Class.BLOCDIVIDENDE = (f"{myresults[0][0]['Equity']} et {myresults[1][0]['Equity']}, la performance positive ou négative de ce placement dépendant de l'évolution de {Class.SJR7} entre {myresults[0][0]['Equity']} ({myresults[0][0]['Dividende']}; code Bloomberg : {myresults[0][0]['Ticker']}; place de cotation : {myresults[0][0]['Sponsor']}; {myresults[0][0]['SiteWeb']}) et {myresults[1][0]['Equity']} ({myresults[1][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[1][0]['Sponsor']}; {myresults[1][0]['SiteWeb']}) et {myresults[2][0]['Equity']} ({myresults[2][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[2][0]['Sponsor']}; {myresults[2][0]['SiteWeb']}) et {myresults[3][0]['Equity']} ({myresults[3][0]['Dividende']}; code Bloomberg : {myresults[3][0]['Ticker']}; place de cotation : {myresults[3][0]['Sponsor']}; {myresults[3][0]['SiteWeb']})")

        elif (len(ticker) == 5):
            Class.BLOCDIVIDENDE = (f"{myresults[0][0]['Equity']} et {myresults[1][0]['Equity']}, la performance positive ou négative de ce placement dépendant de l'évolution de {Class.SJR7} entre {myresults[0][0]['Equity']} ({myresults[0][0]['Dividende']}; code Bloomberg : {myresults[0][0]['Ticker']}; place de cotation : {myresults[0][0]['Sponsor']}; {myresults[0][0]['SiteWeb']}) et {myresults[1][0]['Equity']} ({myresults[1][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[1][0]['Sponsor']}; {myresults[1][0]['SiteWeb']}) et {myresults[2][0]['Equity']} ({myresults[2][0]['Dividende']}; code Bloomberg : {myresults[1][0]['Ticker']}; place de cotation : {myresults[2][0]['Sponsor']}; {myresults[2][0]['SiteWeb']}) et {myresults[3][0]['Equity']} ({myresults[3][0]['Dividende']}; code Bloomberg : {myresults[3][0]['Ticker']}; place de cotation : {myresults[3][0]['Sponsor']}; {myresults[3][0]['SiteWeb']})  et {myresults[4][0]['Equity']} ({myresults[4][0]['Dividende']}; code Bloomberg : {myresults[4][0]['Ticker']}; place de cotation : {myresults[4][0]['Sponsor']}; {myresults[4][0]['SiteWeb']})")

        else:
            Class.BLOCDIVIDENDE  = "errror"

    except Exception:
        Class.BLOCDIVIDENDE = "error"
        print("error")

# add_value()